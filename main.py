import pandas as pd
import sqlalchemy as sa
import os
import openpyxl
from millify import millify

rates_file = "data/SFY 2026 Standard Plan Rate Exhibits w PCs_2025.06.04.xlsx"

def calculate_sfy26_pmpm(base_pmpm,trend,pc,mcs):


    data = {}
    trend_pmpm = base_pmpm * (((1+trend)**(24/12)))
    pc_pmpm = trend_pmpm * (1+pc)
    mcs_pmpm = pc_pmpm * (1+mcs)
    sfy26_pmpm = base_pmpm * ((1+trend)**(24/12)) * (1+pc) * (1+mcs)
    print(f"sfy26_pmpm: {sfy26_pmpm}")

    data['trend_pmpm'] = trend_pmpm
    data['pc_pmpm'] = pc_pmpm
    data['mcs_pmpm'] = mcs_pmpm
    data['mcs_sav_pmpm'] = mcs_pmpm - pc_pmpm

    return data

def aggregate_data(sheet):
    region = sheet["C3"].value
    coa = sheet["C4"].value
    MM = sheet["C7"].value
    base_pmpm = sheet["C42"].value
    trend = sheet["F42"].value
    mcs = sheet["W42"].value
    pc = sheet["I42"].value
    sfy26_pmpm = sheet["X42"].value
    gross_pmpm = sheet["Z46"].value
    service_non_benefit = sheet["Z53"].value
    premium_tax = sheet["Z54"].value
    underwriting_gain = sheet["Z51"].value
    
    cal_sfy26 = sheet
    
    data_dict = calculate_sfy26_pmpm(base_pmpm=base_pmpm,trend=trend,pc=pc,mcs=mcs)

    mcs_sav_pmpm = data_dict['mcs_sav_pmpm']
    cost = sfy26_pmpm * MM
    sav = mcs_sav_pmpm * MM
    gross_cost = gross_pmpm * MM

    non_benefit = gross_pmpm - service_non_benefit
    non_benefit_cost = non_benefit * MM
    premium_tax_cost = premium_tax * MM
    net_cost = cost - non_benefit_cost - premium_tax_cost
    underwriting_cost = underwriting_gain * MM
    

    return {
        "region": region,
        "coa": coa,
        "MM": MM,
        "sfy26_pmpm": sfy26_pmpm,
        "cost": cost,
        "sav": sav,
        "gross_cost": gross_cost,
        "non_benefit_cost": non_benefit_cost,
        "premium_tax_cost": premium_tax_cost,
        "net_cost": net_cost,
        "underwriting_cost": underwriting_cost
    }

def get_data_by_category(sheet):
    data = []
 

    for row in sheet["B16:X41"]:
        row_data = {}
        region = sheet["C3"].value
        coa = sheet["C4"].value
        MM = sheet["C7"].value

        category = row[0].value
        base_pmpm = row[1].value
        trend = row[4].value
        mcs = row[21].value
        pc = row[7].value
        sfy26_pmpm = row[22].value
        gross_pmpm = base_pmpm * ((1+trend)**(24/12)) * (1+pc)
        mcs_sav_pmpm = gross_pmpm * mcs
        
        cost = sfy26_pmpm * MM
        savings = mcs_sav_pmpm * MM

        row_data["region"] = region
        row_data["coa"] = coa
        row_data["MM"] = MM
        row_data["category"] = category
        row_data["base_pmpm"] = base_pmpm
        row_data["trend"] = trend
        row_data["mcs"] = mcs
        row_data["pc"] = pc
        row_data["sfy26_pmpm"] = sfy26_pmpm
        row_data["gross_pmpm"] = gross_pmpm
        row_data["mcs_sav_pmpm"] = mcs_sav_pmpm
        row_data["cost"] = cost
        row_data["savings"] = savings

        data.append(row_data)

    return data

def load_valid_sheets(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    valid_sheets = []
    for sheet in wb.sheetnames:
        if sheet.startswith("R"):
            valid_sheets.append(sheet)
    return valid_sheets

def savings_report(file):
    total_MM = 0
    valid_sheets = load_valid_sheets(file)
    wb = openpyxl.load_workbook(file, data_only=True)

    data_list = []
    total_savings = 0
    grand_total_cost = 0
    total_member_months = 0
    gross_grand_total_cost = 0
    net_grand_total_cost = 0
    non_benefit_grand_total_cost = 0
    premium_tax_grand_total_cost = 0
    underwriting_grand_total_cost = 0

    for sheet in valid_sheets:
        sheet = wb[sheet]
        data = aggregate_data(sheet)
        total_MM += data["MM"]
        total_savings += data["sav"]
        grand_total_cost += data["cost"]
        gross_grand_total_cost += data["gross_cost"]
        net_grand_total_cost += data["net_cost"]
        non_benefit_grand_total_cost += data["non_benefit_cost"]
        premium_tax_grand_total_cost += data["premium_tax_cost"]
        underwriting_grand_total_cost += data["underwriting_cost"]
        data_list.append(data)

    df = pd.DataFrame(data_list)

    agg_df = df.groupby("region").agg({"MM": "sum", "sav": "sum", "cost": "sum"})
    agg_df["sav_pct"] = agg_df["sav"] / agg_df["cost"] * 100
    agg_df["sav_pct"] = agg_df["sav_pct"].apply(lambda x: f"{x:,.2f}%")
    agg_df["MM"] = agg_df["MM"].apply(lambda x: f"{x:,.0f}")
    agg_df["cost"] = agg_df["cost"].apply(lambda x: f"{x:,.2f}")
    agg_df["sav"] = agg_df["sav"].apply(lambda x: f"{x:,.2f}")
    print(agg_df)

    adj_cost = (underwriting_grand_total_cost + total_savings)-premium_tax_grand_total_cost

    print("+"*50)
    print("Total Savings")
    print("+"*50)
    print(f"Total Member Months: {total_MM:,.0f}")
    print(f"Total Members: {total_MM/12:,.0f}")
    print(f"Total Savings: {total_savings:,.2f}")
    print(f"Total Cost: {grand_total_cost:,.2f}")
    print(f"Total Gross Cost: {gross_grand_total_cost:,.2f}")
    print(f"Total Non-Benefit Cost: {non_benefit_grand_total_cost:,.2f}")
    print(f"Total Premium Tax Cost: {premium_tax_grand_total_cost:,.2f}")
    print(f"Total Net Cost: {net_grand_total_cost:,.2f}")
    print(f"Total Underwriting Cost: {underwriting_grand_total_cost:,.2f}")
    print(f"Adjusted Cost: {adj_cost:,.2f}")
    print(f"Total Savings Percentage: {total_savings/grand_total_cost*100:.2f}%")
    print("+"*50)


def savings_by_category(file):
    all_data = []
    valid_sheets = load_valid_sheets(file)
    wb = openpyxl.load_workbook(file, data_only=True)
    for sheet in valid_sheets:
        data = get_data_by_category(wb[sheet])
        all_data.extend(data)
    df = pd.DataFrame(all_data)
    df.to_csv("mcs_saving2.csv")

    return df

def main():
    savings_report(rates_file)
    mcs_savings_df = savings_by_category(rates_file)
    mcs_by_category = mcs_savings_df.groupby("category").agg({"cost": "sum", "savings": "sum"})
    mcs_by_category["cost"] = mcs_by_category["cost"]
    mcs_by_category["savings"] = mcs_by_category["savings"]
    cost = mcs_by_category["cost"].sum()
    savings = mcs_by_category["savings"].sum()
    savings_pct = savings/cost*100


    mcs_by_category = mcs_by_category.sort_values(by="savings", ascending=True)
    display_data = mcs_by_category.copy()
    display_data = display_data.iloc[:3]
    display_data = display_data.reset_index()
    display_data = display_data.set_index("category")
    display_data.to_csv("SFY26_mcs_by_category.csv")

    display_data["savings_pct"] = display_data["savings"]/display_data["cost"]*100
    display_data["savings_pct"] = display_data["savings_pct"].apply(lambda x: f"{x:,.2f}%")
    display_data["cost"] = display_data["cost"].apply(lambda x: millify(x, precision=2))
    display_data["savings"] = display_data["savings"].apply(lambda x: millify(x, precision=2))
    print(display_data)

    print("-"*50)
    print(f"Total Cost: {millify(cost, precision=2)}")
    print(f"Total Savings: {millify(savings, precision=2)}")
    print(f"Total Savings Percentage: {savings_pct:.2f}%")
    print("-"*50)
    

if __name__ == "__main__":
    main()
